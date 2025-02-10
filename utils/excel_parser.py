from enum import Enum

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from model.main_db.teacher import TeacherRaw
from model.main_db.student import StudentRaw

class ExcelDataParserError(Exception):
    pass

class ParserType(Enum):
    ALL = 0
    TEACHER = 1
    STUDENT = 2

class ExcelDataParser:
    def __init__(self, file_path: str, parse_type: ParserType = ParserType.ALL):
        self.__student: dict[str, dict[str, list[StudentRaw]]] = {}
        self.__teacher: dict[str, dict[str, list[TeacherRaw]]] = {}
        self.__parse_type = parse_type
        self.__load_data(file_path, parse_type)

    @property
    def students(self) -> dict[str, dict[str, list[StudentRaw]]]:
        if self.__parse_type == ParserType.TEACHER:
            raise ExcelDataParserError("Students data don't with this ParseType")
        return self.__student

    @property
    def teachers(self) -> dict[str, dict[str, list[TeacherRaw]]]:
        if self.__parse_type == ParserType.TEACHER:
            raise ExcelDataParserError("Teachers data don't with this ParseType")
        return self.__teacher


    def __load_data(self, file_path: str, parse_type: ParserType) -> None:
        wb: Workbook = openpyxl.load_workbook(file_path)

        match parse_type:
            case ParserType.ALL:
                index = wb.sheetnames.index('teachers')
                wb.active = index
                self.__students_parser(wb.active)
                index = wb.sheetnames.index('students')
                wb.active = index
                self.__teachers_parser(wb.active)
            case ParserType.STUDENT:
                index = wb.sheetnames.index('students')
                wb.active = index
                self.__teachers_parser(wb.active)
            case ParserType.TEACHER:
                index = wb.sheetnames.index('teachers')
                wb.active = index
                self.__students_parser(wb.active)
            case _:
                raise ExcelDataParserError('ParserType not found')


    def __teachers_parser(self, worksheet: Worksheet) -> None:
        teacher_name = ''
        row = 2

        while teacher_name is not None:
            worksheet_list = [worksheet.cell(row=row, column=i).value for i in range(0,7)]
            teacher_name = worksheet_list[1]
            telegram_id = worksheet_list[2]
            discipline = worksheet_list[3]
            is_admin = bool(worksheet_list[4])
            group =  worksheet_list[5]

            if teacher_name is None:
                break

            if discipline not in self.__teacher:
                self.__teacher[discipline] = {}

            if group not in self.__teacher[discipline]:
                self.__teacher[discipline][group] = []

            self.__teacher[discipline][group].append(
                TeacherRaw(
                    full_name=teacher_name,
                    telegram_id=telegram_id,
                    is_admin=is_admin
                )
            )

            row += 1

    def __students_parser(self, worksheet: Worksheet) -> None:
        student_name = ''
        row = 2

        while student_name is not None:
            worksheet_list = [worksheet.cell(row=row, column=i).value for i in range(0,4)]
            student_name = worksheet_list[1]
            group = worksheet_list[2]
            discipline = worksheet_list[3]

            if student_name is None:
                break

            if discipline not in self.__student:
                self.__student[discipline] = {}

            if group not in self.__student[discipline]:
                self.__student[discipline][group] = []

            self.__student[discipline][group].append(
                StudentRaw(
                    full_name=student_name
                )
            )

            row += 1
